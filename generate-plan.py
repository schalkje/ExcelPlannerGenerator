import os
# https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
from openpyxl.styles.protection import Protection
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
# from openpyxl.workbook.protection import WorkbookProtection
from copy import copy
import datetime
import calendar


# https://www.blog.pythonlibrary.org/2021/08/11/styling-excel-cells-with-openpyxl-and-python/#:~:text=Adding%20a%20Border%20OpenPyXL%20gives%20you%20the%20ability,each%20of%20the%20four%20sides%20of%20a%20cell.


# input variables
startDate = datetime.date(2022, 1, 1)
numberOfMonths = 13
first_flip_day = datetime.date(2022, 1, 11)
sprint_size = 14 # (2 weeks)

first_sprint_nr = 1
teamMembersLabels = ['Name','Start','End','Monday','Tuesday','Wednessday','Thursday','Friday']
teamMembers = [
                    ['David',datetime.date(2022, 1, 1),None,8,8,8,8,8],
                    ['Maria',datetime.date(2022, 1, 1),None,8,8,8,8,8],
                    ['Gijs',datetime.date(2022, 1, 1),None,8,8,8,8,8],
                    ['Bart',datetime.date(2022, 1, 1),None,8,8,8,8,8],
                    ['Paul',datetime.date(2022, 1, 1),datetime.date(2022, 3, 31),8,0,0,8,0],
                    ['Timothy',datetime.date(2022, 3, 1),None,8,8,8,8,0],
                    ['Michel',datetime.date(2022, 1, 1),None,8,8,8,8,8],
                    ['Jeroen',datetime.date(2022, 1, 1),None,-8,-8,0,-8,-8],
                ]

# computed variables
num_sprints = int((numberOfMonths*30)/sprint_size)+1
flip_days = [first_flip_day + datetime.timedelta(days=(x*sprint_size)) for x in range(-1,num_sprints)]
sprint_year = first_flip_day.year

# create a workbook
wb = Workbook()
wb.security.workbookPassword = 'password' # a simple basic password to prevent unconscious changes
wb.security.lockStructure = True

# Styles
# thin = Side(border_style="thin", color="666666")
# thin_inactive = Side(border_style="thin", color="AAAAAA")
thin = Side(border_style="medium", color="FFFFFF")
thick = Side(border_style="thick", color="FFFFFF")
thin_flip = Side(border_style="medium", color="DFE18F")
thin_alternate = Side(border_style="medium", color="808080")
thin_inactive = Side(border_style="medium", color="FFFFFF")
double = Side(border_style="double", color="666666")
side_team = Side(border_style="thick", color="FFFFFF")
side_team_vertical = Side(border_style="medium", color="FFFFFF")

fill_inactive = PatternFill("solid", fgColor="00EEEEEE")
fill_team_inactive = PatternFill("solid", fgColor="00FFFFFF")
fill_active_header = PatternFill("solid", fgColor="00B7D2FF")
fill_header_label = PatternFill("solid", fgColor="FFFFFF")
fill_weekend = PatternFill("solid", fgColor="00FFFFFF")
fill_not_working = PatternFill("solid", bgColor="00FFFFFF")
fill_workday_odd = PatternFill("solid", fgColor="00DCE6F1")
fill_workday_even = PatternFill("solid", fgColor="00B8CCE4")
fill_footer = PatternFill("solid", fgColor="00DFE18F")
fill_footer_inactive = PatternFill("solid", fgColor="00EEEEEE")
fill_footer_alternate = PatternFill("solid", fgColor="00808080")
fill_flip_day= PatternFill("solid", fgColor="00DFE18F")

font_inactive = Font(color="999999", bold=False)

style_team = NamedStyle(name="team")

style_day = NamedStyle(name="day")
style_day.font = Font(color="000000", bold=True)
style_day.fill = fill_active_header
style_day.alignment = Alignment(horizontal="center", vertical="center")
style_day.border = Border(top=thin, left=thin, right=thin, bottom=thin)


sick_font = Font(color="DA9694", bold=False)
sick_fill = PatternFill("solid", bgColor="00FFD5D5")
vacation_font = Font(color="C4D79B", bold=False)
vacation_fill = PatternFill("solid", bgColor="0076933C")
not_contributing_font = Font(color="FFFFFF", bold=False)
not_working_font = Font(color="F9F9F9", bold=False)
# style_sick_day = copy(style_day)
# style_sick_day.name = "sick_day"
# style_sick_day.font = Font(color="DA9694", bold=False)
# style_sick_day.fill = PatternFill("solid", fgColor="00FFD5D5")



style_flip_day = NamedStyle(name="flip_day")
style_flip_day.font = Font(color="000000", bold=True)
style_flip_day.fill = fill_flip_day
style_flip_day.alignment = Alignment(horizontal="center", vertical="center")
style_flip_day.border = Border(top=thin, left=thin, right=thin, bottom=thin)


style_day_inactive = copy(style_day)
style_day_inactive.name = "day_inactive"
style_day_inactive.font = font_inactive
style_day_inactive.fill = fill_inactive
style_day_inactive.border = Border(top=thin_inactive, left=thin_inactive, right=thin_inactive, bottom=thin_inactive)

style_weekday = NamedStyle(name="weekday")
style_weekday.alignment = Alignment(horizontal="center", vertical="center")
style_weekday.border = Border(top=thin, left=thin, right=thin, bottom=thin)
style_weekday.fill = fill_active_header

style_weekday_inactive = copy(style_day)
style_weekday_inactive.name = "weekday_inactive"
style_weekday_inactive.font = font_inactive
style_weekday_inactive.fill = fill_inactive
style_weekday_inactive.border = Border(top=thin_inactive, left=thin_inactive, right=thin_inactive, bottom=thin_inactive)

style_weekend = NamedStyle(name="weekend")
style_weekend.alignment = Alignment(horizontal="center", vertical="center")
style_weekend.border = Border(bottom=thin)
style_weekend.font = font_inactive
style_weekend.fill = fill_weekend

style_weekend_inactive = copy(style_weekend)
style_weekend_inactive.name = "weekend_inactive"
style_weekend_inactive.font = font_inactive
style_weekend_inactive.fill = fill_weekend

# rule = Rule()

style_workday_odd = NamedStyle(name="workday_odd")
style_workday_odd.alignment = Alignment(horizontal="center", vertical="center")
style_workday_odd.border = Border(bottom=side_team, left=side_team_vertical, right=side_team_vertical)
style_workday_odd.fill = fill_workday_odd

style_team_odd = copy(style_workday_odd)
style_team_odd.name = "style_team_odd"
style_team_odd.alignment = Alignment(horizontal="left", vertical="center")
style_team_odd.font = Font(bold=True)

style_workday_even = NamedStyle(name="workday_even")
style_workday_even.alignment = Alignment(horizontal="center", vertical="center")
style_workday_even.border = Border(bottom=side_team, left=side_team_vertical, right=side_team_vertical)
style_workday_even.fill = fill_workday_even

style_team_even = copy(style_workday_even)
style_team_even.name = "style_team_even"
style_team_even.alignment = Alignment(horizontal="left", vertical="center")
style_team_even.font = Font(bold=True)

style_not_working = NamedStyle(name="not_working")
style_not_working.alignment = Alignment(horizontal="center", vertical="center")
style_not_working.border = Border(bottom=thin)
style_not_working.font = not_working_font
style_not_working.fill = fill_weekend

style_workday_inactive = copy(style_day)
style_workday_inactive.name = "workday_inactive"
style_workday_inactive.font = font_inactive
style_workday_inactive.fill = fill_inactive
style_workday_inactive.border = Border(bottom=side_team, left=side_team_vertical, right=side_team_vertical)


style_month = NamedStyle(name="month")
style_month.font = Font(color="000000", bold=True)
style_month.border = Border(top=thin, left=thin, right=thin, bottom=thin)
style_month.alignment = Alignment(horizontal="center", vertical="center")
style_month.fill = fill_active_header

style_month_inactive = copy(style_month)
style_month_inactive.name = 'month_inactive'
style_month_inactive.font = font_inactive
style_month_inactive.fill = fill_inactive
style_month_inactive.border = Border(top=thin_inactive, left=thin_inactive, right=thin_inactive, bottom=thin_inactive)

style_team_header = NamedStyle(name="style_team_header")
style_team_header.font = Font(color="000000", bold=True)
# style_team_header.border = Border(top=thin, left=thin, right=thin, bottom=thin)
style_team_header.alignment = Alignment(horizontal="center", vertical="bottom")
style_team_header.fill = fill_header_label

style_team = NamedStyle(name="style_team")
style_team.font = Font(color="000000", bold=False)
style_team.border = Border(top=thin, left=thin, bottom=thin)
style_team.alignment = Alignment(horizontal="left", vertical="center")
style_team.fill = fill_active_header

style_team_inactive = copy(style_team)
style_team_inactive.name = "style_team_inactive"
style_team_inactive.font = font_inactive
style_team_inactive.border = Border(top=thin_inactive, left=thin_inactive, bottom=thin_inactive)
style_team_inactive.alignment = Alignment(horizontal="left", vertical="center")
style_team_inactive.fill = fill_team_inactive

style_team_workday_inactive = copy(style_team)
style_team_workday_inactive.name = "style_team_workday_inactive"
style_team_workday_inactive.font = font_inactive
style_team_workday_inactive.border = Border(top=thin_inactive, left=thin_inactive, bottom=thin_inactive)
style_team_workday_inactive.alignment = Alignment(horizontal="center", vertical="center")
style_team_workday_inactive.fill = fill_team_inactive


style_footer_sum = NamedStyle(name="footer_sum")
style_footer_sum.font = Font(color="808080", bold=False)
style_footer_sum.border = Border(top=thin, left=thin, right=thin)
style_footer_sum.alignment = Alignment(horizontal="center", vertical="center")
style_footer_sum.fill = fill_footer

style_footer_sum_inactive = copy(style_footer_sum)
style_footer_sum_inactive.name="footer_sum_inactive"
style_footer_sum_inactive.fill = fill_footer_inactive

style_footer_sprint = NamedStyle(name="footer_sprint")
style_footer_sprint.font = Font(color="808080", bold=False)
style_footer_sprint.fill = fill_footer
style_footer_sprint.alignment = Alignment(horizontal="center", vertical="center")
style_footer_sprint.border = Border(left=thin, right=thin_flip)


style_footer_line = NamedStyle(name="footer_line")
style_footer_line.font = Font(color="000000", bold=True)
style_footer_line.fill = fill_footer_alternate
style_footer_line.alignment = Alignment(horizontal="right", vertical="center")
style_footer_line.border = Border(left=thin)

style_footer_sprint_total = NamedStyle(name="footer_sprint_total")
style_footer_sprint_total.font = Font(color="FFFFFF", bold=True)
style_footer_sprint_total.fill = fill_footer_alternate
style_footer_sprint_total.alignment = Alignment(horizontal="center", vertical="center")
style_footer_sprint_total.border = Border(right=thick) # , left=thin_alternate


def addMonths(d, x):
    newday = d.day
    newmonth = (((d.month - 1) + x) % 12) + 1
    newyear  = d.year + (((d.month - 1) + x) // 12)
    if newday > calendar.mdays[newmonth]:
        newday = calendar.mdays[newmonth]
        if newyear % 4 == 0 and newmonth == 2:
            newday += 1
    return datetime.date(newyear, newmonth, newday)





generatorVariables = [
    ['startDate',startDate],
    ['numberOfMonths',numberOfMonths],
]


# grab the active worksheet and rename
ws_overview = wb.active
ws_overview.title = "Overview"
props = ws_overview.sheet_properties
props.tabColor = "1072BA"

ws_overview.append(['Variable', 'Value'])
ws_overview.column_dimensions['A'].width = 22
ws_overview.column_dimensions['B'].width = 30

for row in generatorVariables:
    ws_overview.append(row)

tab = Table(displayName="GeneratorVariables", ref="A1:B3")
# Add a default style with striped rows_overview and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

ws_overview.add_table(tab)

# empty line
ws_overview.append([''])

startRowNumber = 5
ws_overview.append(teamMembersLabels)
ws_overview.column_dimensions['A'].width = 22
ws_overview.column_dimensions['B'].width = 30
ws_overview.column_dimensions['C'].width = 30
ws_overview.column_dimensions['D'].width = 10
ws_overview.column_dimensions['E'].width = 10
ws_overview.column_dimensions['F'].width = 10
ws_overview.column_dimensions['G'].width = 10
ws_overview.column_dimensions['H'].width = 10

for row in teamMembers:
    ws_overview.append(row)

tab = Table(displayName="TeamMembers", ref="A{0}:H{1}".format(startRowNumber,startRowNumber+len(teamMembers)))
# Add a default style with striped rows_overview and banded columns
style = TableStyleInfo(name="TableStyleMedium11", showFirstColumn=True,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style

ws_overview.add_table(tab)



# creating the month sheets
sprint_nr = first_sprint_nr -1 # start one lower, because the iteration starts with incrementing this counter
sprint_already_incremented = False
sprint_label = "Previous sprint"
sheet_title = ""
endMonthDays = 0
previous_endMonthDays = 0
days_in_current_month = 0

for monthNumber in range(0,numberOfMonths):
    date = addMonths(startDate,monthNumber)

    previous_sheet_title = addMonths(date,-1).strftime('%Y-%m')
    sheet_title = date.strftime('%Y-%m')
    next_sheet_title = addMonths(date,1).strftime('%Y-%m')

    ws = wb.create_sheet(title=sheet_title)

    
    # props = ws.sheet_properties
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showRowColHeaders = False 


        
    cal = calendar.Calendar(0)
    days = cal.itermonthdates(date.year, date.month)

    offset_rows = 1
    offset_cols = 3
    team_offset_cols = 2
    team_offset_rows = 5    

    startHeaderDays = 1
    endHeaderDays = date.weekday()
    endHeaderDays_next_month = addMonths(date,1).weekday()
    previous_endMonthDays = endMonthDays
    days_in_previous_month = days_in_current_month
    days_in_current_month = calendar.monthrange(date.year,date.month)[1]
    endMonthDays = date.weekday()+days_in_current_month
    endTailDates = (int( (endMonthDays) / 7 ) + 1) * 7 
    column_nr = offset_cols
    day_counter = 0
    for day in days:
        column_nr += 1
        day_counter += 1
        col = get_column_letter(column_nr)
        ws['{0}{1}'.format(col,2)]=day.strftime('%b') # Month as locale’s abbreviated name.  (https://strftime.org/)
        ws['{0}{1}'.format(col,4)]=day.strftime('%a') # Weekday as locale’s abbreviated name. (https://strftime.org/)

        ws['{0}{1}'.format(col,3)]=day.day
        if ( day_counter > endHeaderDays and day_counter <= endMonthDays ):
            ws['{0}{1}'.format(col,3)].style = style_day
            ws['{0}{1}'.format(col,2)].style = style_month
            if day.strftime('%w') == "0" or day.strftime('%w') == "6":
                ws['{0}{1}'.format(col,4)].style = style_weekend
            else:
                ws['{0}{1}'.format(col,4)].style = style_weekday
        else:
            ws['{0}{1}'.format(col,3)].style = style_day_inactive
            ws['{0}{1}'.format(col,2)].style = style_month_inactive
            if day.strftime('%w') == "0" or day.strftime('%w') == "6":
                ws['{0}{1}'.format(col,4)].style = style_weekend_inactive
            else:
                ws['{0}{1}'.format(col,4)].style = style_weekday_inactive

        ws.column_dimensions[col].width = 5

    # merge month header 
    if (endHeaderDays > startHeaderDays):
        ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(offset_cols + startHeaderDays),2,get_column_letter(offset_cols + endHeaderDays),2))
    
    ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(offset_cols + endHeaderDays + 1),2,get_column_letter(offset_cols + endMonthDays),2))
    # 1 + date.weekday()+calendar.monthrange(date.year,date.month)[1]
    if (endMonthDays < endTailDates):
        ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(offset_cols + endMonthDays + 1),2,get_column_letter(offset_cols + endTailDates),2))


    # Team member lines


    ws['{}{}'.format(get_column_letter(team_offset_cols),team_offset_rows-1)]='Team'
    ws['{}{}'.format(get_column_letter(team_offset_cols),team_offset_rows-1)].style = style_team_header
    ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(team_offset_cols),team_offset_rows-1,get_column_letter(team_offset_cols),team_offset_rows))
    ws.row_dimensions[team_offset_rows].height = 4

    # add empty columns
    for i in range(1,team_offset_cols):
        ws.column_dimensions[get_column_letter(i)].width = 4
    # add empty columns
    for i in range(team_offset_cols+1,offset_cols+1):
        ws.column_dimensions[get_column_letter(i)].width = 1

    counter = 0
    for teamMember in teamMembers:
        counter += 1
        column_nr = team_offset_cols

        row = team_offset_rows + counter
        col = get_column_letter(column_nr)
        
        ws['{0}{1}'.format(col,row)]=teamMember[0]
        if (teamMember[1] is None or teamMember[1] <= date) and (teamMember[2] is None or teamMember[2] > date):
            if (counter % 2) == 0:
                ws['{0}{1}'.format(col,row)].style = style_team_even
            else:
                ws['{0}{1}'.format(col,row)].style = style_team_odd
        else:
            ws['{0}{1}'.format(col,row)].style = style_team_inactive

        


        column_nr = offset_cols
        day_counter = 0

        days = cal.itermonthdates(date.year, date.month)

        for day in days:
            column_nr += 1
            day_counter += 1

            col = get_column_letter(column_nr)

            if (teamMember[1] is None or teamMember[1] <= date) and (teamMember[2] is None or teamMember[2] > date):
                if ( day_counter > endHeaderDays and day_counter <= endMonthDays  ):
                    # weekend days are non-working days and greyed out as such
                    if day.strftime('%w') == "0" or day.strftime('%w') == "6":
                        ws['{0}{1}'.format(col,row)]='="o"' 
                        ws['{0}{1}'.format(col,row)].style=style_weekend
                    else:
                        # check if sprint flip day - style_flip_day
                        if (day in flip_days):
                            ws['{0}{1}'.format(col,row)].style=style_flip_day
                        else:
                            if (counter % 2) == 0:
                                ws['{0}{1}'.format(col,row)].style=style_workday_even
                            else:
                                ws['{0}{1}'.format(col,row)].style=style_workday_odd

                            if (teamMember[2 + int(day.strftime('%w'))]==0):
                                ws['{0}{1}'.format(col,row)] = '="o"'
                            else:
                                if (teamMember[2 + int(day.strftime('%w'))]<0):
                                    ws['{0}{1}'.format(col,row)] = 'x'

                else:
                    # inactive days are inactive and copy the value from the previous month 

                    if (day in flip_days):
                        ws['{0}{1}'.format(col,row)].style=style_flip_day
                    else:
                        # inactive days are locked
                        if day.strftime('%w') == "0" or day.strftime('%w') == "6":
                            ws['{0}{1}'.format(col,row)]='="o"'
                            ws['{0}{1}'.format(col,row)].style=style_weekend_inactive
                        else:
                            ws['{0}{1}'.format(col,row)].style=style_workday_inactive
                            if previous_endMonthDays == 0:
                                ws['{0}{1}'.format(col,row)]='="o"'
                            
                        if previous_endMonthDays > 0 and day_counter < 10: # looking back
                            parallel_col_nr = offset_cols + previous_endMonthDays - (days_in_previous_month - day.day)
                            ws['{0}{1}'.format(col,row)]="=IF('{0}'!{1}{2}=\"\",\"\",'{0}'!{1}{2})".format(previous_sheet_title,get_column_letter(parallel_col_nr),row)
                        if monthNumber < numberOfMonths-1 and day_counter > 10: # looking forward
                            ws['{0}{1}'.format(col,row)].style=style_workday_inactive                      
                            parallel_col_nr = offset_cols + endHeaderDays_next_month + day.day
                            ws['{0}{1}'.format(col,row)]="=IF('{0}'!{1}{2}=\"\",\"\",'{0}'!{1}{2})".format(next_sheet_title,get_column_letter(parallel_col_nr),row)

            else:
                if (day in flip_days):
                    ws['{0}{1}'.format(col,row)].style=style_flip_day
                else:
                    ws['{0}{1}'.format(col,row)]='="o"'
                    if ( day_counter > endHeaderDays and day_counter <= endMonthDays  ):
                        ws['{0}{1}'.format(col,row)]='="o"'
                        ws['{0}{1}'.format(col,row)].style=style_team_workday_inactive
                    else:
                        if previous_endMonthDays > 0 and day_counter < 10: # looking back
                            parallel_col_nr = offset_cols + previous_endMonthDays - (days_in_previous_month - day.day)
                            ws['{0}{1}'.format(col,row)]="=IF('{0}'!{1}{2}=\"\",\"\",'{0}'!{1}{2})".format(previous_sheet_title,get_column_letter(parallel_col_nr),row)
                        if monthNumber < numberOfMonths-1 and day_counter > 10: # looking forward
                            ws['{0}{1}'.format(col,row)].style=style_workday_inactive                      
                            parallel_col_nr = offset_cols + endHeaderDays_next_month + day.day
                            ws['{0}{1}'.format(col,row)]="=IF('{0}'!{1}{2}=\"\",\"\",'{0}'!{1}{2})".format(next_sheet_title,get_column_letter(parallel_col_nr),row)



    # summary footer
    column_nr = offset_cols
    day_counter = 0

    days = cal.itermonthdates(date.year, date.month)

    offset_footer = 3
    row = team_offset_rows + counter + offset_footer
    col_sprint_start = offset_footer+1

    for day in days:
        column_nr += 1
        day_counter += 1

        col = get_column_letter(column_nr)

        if day.strftime('%w') == "0" or day.strftime('%w') == "6":
            ws['{0}{1}'.format(col,row)].style=style_weekend
        else:
            if (day in flip_days):
                formula_current_month = "SUM({}{}:{}{})".format(get_column_letter(col_sprint_start),row,get_column_letter(column_nr-1),row)
                
                if day_counter < sprint_size and previous_endMonthDays > 0:
                    if day.day < 10:
                        parallel_start_col = offset_cols + previous_endMonthDays - (14 - day.day - 1)
                        parallel_end_col = offset_cols + previous_endMonthDays - (day_counter - day.day) + 1
                        formula_previous_month = "SUM('{0}'!{1}{2}:{3}{4})".format(previous_sheet_title,get_column_letter(parallel_start_col),row,get_column_letter(parallel_end_col-1),row)
                    else: # sprint flip previous month
                        parallel_start_col = offset_cols + previous_endMonthDays - (14 - day_counter) - 1
                        parallel_end_col = offset_cols + previous_endMonthDays - (day_counter) + 1
                        formula_previous_month = "SUM('{0}'!{1}{2}:{3}{4})".format(previous_sheet_title,get_column_letter(parallel_start_col),row,get_column_letter(parallel_end_col-1),row)

                    ws['{0}{1}'.format(col,row)] = "={}+{}".format(formula_current_month,formula_previous_month)
                else:
                    ws['{0}{1}'.format(col,row)] = "={}".format(formula_current_month)
                ws['{0}{1}'.format(col,row)].style=style_footer_sprint_total
                ws.merge_cells('{0}{1}:{2}{3}'.format(col,row,col,row+2))

                # sprint numbering resets at the start of the year
                if ( day_counter > endHeaderDays and day_counter <= endMonthDays  ):
                    if sprint_already_incremented:
                        sprint_already_incremented = False
                    else:                    
                        if sprint_year != day.year:
                            sprint_year = day.year
                            sprint_nr = 1
                        else:
                            sprint_nr += 1
                    # sprint name
                    sprint_label = "Sprint {}-{:02d}".format(day.year,sprint_nr)
                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].style=style_footer_sprint
                    if (column_nr <= sprint_size) and (column_nr - col_sprint_start) < 4:
                        ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].alignment = Alignment(horizontal="right", vertical="center")
                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)]=sprint_label
                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+2)].style=style_footer_line
                else:
                    # previous sprint name in case a sprint flip in the header
                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].style=style_footer_sprint
                    if (column_nr <= sprint_size) and (column_nr - col_sprint_start) < 4:
                        ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        if (column_nr - col_sprint_start) < 4:
                            ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].alignment = Alignment(horizontal="left", vertical="center")

                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)]=sprint_label #TODO  error !? 2022-11
                    ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+2)].style=style_footer_line

                # merge sprint name
                if col_sprint_start < column_nr-1: # skip merge when there is nothing to merge
                    ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(col_sprint_start),row+1,get_column_letter(column_nr-1),row+1))
                    ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(col_sprint_start),row+2,get_column_letter(column_nr-1),row+2))
                col_sprint_start = column_nr+1
                
            else:
                # sum day availability
                ws['{0}{1}'.format(col,row)]='=COUNTIF({}{}:{}{},"")'.format(col,team_offset_rows+1,col,team_offset_rows+len(teamMembers))
                if ( day_counter > endHeaderDays and day_counter <= endMonthDays  ):
                    ws['{0}{1}'.format(col,row)].style=style_footer_sum
                else:
                    ws['{0}{1}'.format(col,row)].style=style_footer_sum_inactive

    if col_sprint_start < column_nr:
        if sprint_year != (day+datetime.timedelta(days=sprint_size - (column_nr - col_sprint_start))).year:
            sprint_year = (day+datetime.timedelta(days=sprint_size - (column_nr - col_sprint_start))).year
            sprint_nr = 1
        else:
            sprint_nr += 1        # sprint name
        sprint_already_incremented = True
        ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].style=style_footer_sprint
        if (column_nr - col_sprint_start) < 4:
            ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)].alignment = Alignment(horizontal="left", vertical="center")

        ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+1)]="Sprint {}-{:02d}".format(day.year,sprint_nr)
        ws['{0}{1}'.format(get_column_letter(col_sprint_start),row+2)].style=style_footer_line        # merge sprint name
        if col_sprint_start < column_nr-1: # skip merge when there is nothing to merge
            ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(col_sprint_start),row+1,get_column_letter(column_nr),row+1))
            ws.merge_cells('{0}{1}:{2}{3}'.format(get_column_letter(col_sprint_start),row+2,get_column_letter(column_nr),row+2))
    ws.row_dimensions[row+2].height = 4

    ws.conditional_formatting.add('A1:BB50', CellIsRule(operator='equal', formula=['="S"'], stopIfTrue=True, fill=sick_fill, font=sick_font))
    ws.conditional_formatting.add('A1:BB50', CellIsRule(operator='equal', formula=['="V"'], stopIfTrue=True, fill=vacation_fill, font=vacation_font))
    ws.conditional_formatting.add('A1:BB50', CellIsRule(operator='equal', formula=['="x"'], stopIfTrue=True, font=not_contributing_font))
    ws.conditional_formatting.add('A1:BB50', CellIsRule(operator='equal', formula=['="o"'], stopIfTrue=True, fill=fill_not_working, font=not_working_font))

    # protect the sheet against unwanted changes
    ws.protection.sheet = True
    ws.protection.enable()
    for c in range( offset_cols+1,offset_cols+day_counter+1):
        for r in range(team_offset_rows+1,team_offset_rows+len(teamMembers)+1):
            ws['{0}{1}'.format(get_column_letter(c),r)].protection = Protection(locked=False)



# Save the file
filename = "output\plan.xlsx"
wb.save(filename)

# open the Excel file
os.system("{}".format(filename))